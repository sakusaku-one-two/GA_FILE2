import os
from abc import abstractclassmethod,ABCMeta

import openpyxl as op 



def new_templates():
    """テンプレを返す"""
    wb_name = '/健康診断＿予約枠自動当て込み.xlsx'
    out_path = os.getcwd()
    new_wb = op.Workbook()
    emp_sheet = new_wb.create_sheet('社員データ')
    reserved_day_sheet = new_wb.create_sheet('予約枠')
    
    #社員シート
    set_list =[
        '社員番号','氏名',
        '病院名','病院番号','検診コース','胃検診',
        '希望日1','希望日2','希望日3','希望時間',
    ]
    i = 1
    for col in set_list:
        emp_sheet.cell(1,i).value = col
        i +=1
        
    
    #予約枠シート
    set_list = [
        '病院名','病院番号','検診コース','胃検診',
        '予約日','予約時間',
    ]
    i = 1
    for col in set_list:
        reserved_day_sheet.cell(1,i).value = col
        i += 1
        
    new_wb.save(out_path+wb_name)




class IState(Meta = ABCMeta):

    @abstractclassmethod
    def __init__(self,contlorer) -> None:
        
    
    @abstractclassmethod
    def do(self):
        pass
   
   
    
class Load(IState):
    
    def __init__(self,controler):
        self.controller = controler
    
    
    def do(self,):pass
    
    
    
        